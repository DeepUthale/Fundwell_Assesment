"""Excel report generator — produces styled .xlsx reports from processed applications."""

from __future__ import annotations

import os
from datetime import datetime, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import EXPORT_DIR
from src.models import LoanApplicationState

# Column definitions: (header_name, field_key, width, number_format)

COLUMNS = [
    ("Date Received", "date_received", 20, None),
    ("Business Name", "business_name", 25, None),
    ("Owner Name", "owner_name", 20, None),
    ("Email", "sender", 30, None),
    ("Location", "location", 20, None),
    ("Years in Business", "years_in_business", 18, "0"),
    ("Monthly Revenue ($)", "monthly_revenue", 20, "#,##0"),
    ("Loan Amount ($)", "loan_amount_requested", 20, "#,##0"),
    ("Loan Purpose", "loan_purpose", 35, None),
    ("Existing Debt", "existing_debt", 25, None),
    ("Revenue/Loan Ratio", "revenue_to_loan_ratio", 18, "0.00"),
    ("Risk Flags", "risk_flags", 45, None),
    ("Confidence Score", "confidence_score", 16, "0.00"),
    ("Triage Decision", "triage_decision", 18, None),
    ("Assigned Underwriter", "assigned_underwriter", 22, None),
    ("Summary", "summary", 55, None),
]

# Styles

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
HEADER_FILL = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DECISION_FILLS = {
    "qualified": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    "needs_review": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    "rejected": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

BODY_FONT = Font(size=10, name="Calibri")
BODY_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

# Summary sheet styles
STAT_LABEL_FONT = Font(bold=True, size=11, name="Calibri")
STAT_VALUE_FONT = Font(size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, name="Calibri", color="2B4C7E")


def _get_cell_value(app: LoanApplicationState, field_key: str):
    """Extract display value from application state."""
    value = app.get(field_key)

    if field_key == "risk_flags" and isinstance(value, list):
        return "; ".join(value) if value else "None"

    if value is None:
        return "N/A"

    return value


def generate_report(
    applications: list[LoanApplicationState],
    filename: Optional[str] = None,
) -> str:
    """Generate a styled Excel report from processed loan applications.

    Args:
        applications: List of processed application states.
        filename: Optional output filename. Auto-generated if not provided.

    Returns:
        Absolute path to the generated .xlsx file.
    """
    os.makedirs(EXPORT_DIR, exist_ok=True)

    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"loan_triage_report_{timestamp}.xlsx"

    filepath = os.path.join(EXPORT_DIR, filename)

    wb = openpyxl.Workbook()

    
    # Sheet 1: Triage Report (detailed)
    ws = wb.active
    ws.title = "Triage Report"

    # Header row
    for col_idx, (header, _, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, app in enumerate(applications, 2):
        decision = app.get("triage_decision", "")
        row_fill = DECISION_FILLS.get(decision)

        for col_idx, (_, field_key, _, num_fmt) in enumerate(COLUMNS, 1):
            value = _get_cell_value(app, field_key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER

            if row_fill:
                cell.fill = row_fill
            if num_fmt and value != "N/A":
                cell.number_format = num_fmt

    # Freeze header row + enable auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Set row height for header
    ws.row_dimensions[1].height = 30

   
    # Sheet 2: Summary dashboard
    summary_ws = wb.create_sheet("Summary")

    # Title
    title_cell = summary_ws.cell(row=1, column=1, value="Loan Triage Summary Report")
    title_cell.font = TITLE_FONT

    summary_ws.merge_cells("A1:C1")

    total = len(applications)
    qualified = sum(1 for a in applications if a.get("triage_decision") == "qualified")
    needs_review = sum(1 for a in applications if a.get("triage_decision") == "needs_review")
    rejected = sum(1 for a in applications if a.get("triage_decision") == "rejected")
    total_requested = sum(a.get("loan_amount_requested", 0) or 0 for a in applications)
    avg_confidence = (
        sum(a.get("confidence_score", 0) or 0 for a in applications) / total
        if total > 0
        else 0
    )

    stats = [
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Applications", total),
        ("Qualified", qualified),
        ("Needs Review", needs_review),
        ("Rejected", rejected),
        ("", ""),
        ("Total Loan Amount Requested", f"${total_requested:,.0f}"),
        ("Average Confidence Score", f"{avg_confidence:.2f}"),
    ]

    # Add qualified rate
    if total > 0:
        stats.append(("Qualification Rate", f"{qualified / total * 100:.1f}%"))

    for row_idx, (label, value) in enumerate(stats, 3):
        label_cell = summary_ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = STAT_LABEL_FONT

        value_cell = summary_ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = STAT_VALUE_FONT

    summary_ws.column_dimensions["A"].width = 32
    summary_ws.column_dimensions["B"].width = 25

    # Color-code the decision counts
    decision_rows = {"Qualified": 6, "Needs Review": 7, "Rejected": 8}
    for label, row in decision_rows.items():
        fill = DECISION_FILLS.get(label.lower().replace(" ", "_"))
        if fill:
            for col in range(1, 3):
                summary_ws.cell(row=row, column=col).fill = fill

    # Save
    wb.save(filepath)
    print(f"[excel] Report saved to {filepath}")
    return os.path.abspath(filepath)


def export_applications(
    applications: list[LoanApplicationState],
    date_range: str = "today",
    filename: Optional[str] = None,
) -> str:
    """Filter applications by date range and generate a report.

    Args:
        applications: All processed applications.
        date_range: "today", "week", or "all".
        filename: Optional output filename.

    Returns:
        Path to the generated Excel file.
    """
    now = datetime.now()

    if date_range == "today":
        cutoff = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    elif date_range == "week":
        cutoff = (now - timedelta(days=7)).isoformat()
    else:
        cutoff = ""

    if cutoff:
        filtered = [
            a
            for a in applications
            if (a.get("date_received") or "") >= cutoff
        ]
    else:
        filtered = list(applications)

    if not filtered:
        print(f"[excel] No applications found for range '{date_range}'.")
        return ""

    return generate_report(filtered, filename)
